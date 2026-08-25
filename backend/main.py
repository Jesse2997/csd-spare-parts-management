from __future__ import annotations

import csv
import io
import json
import re
import sqlite3
import sys
import zipfile
from collections import Counter
from datetime import date, datetime
from pathlib import Path
from typing import Any

from fastapi import File, Form, HTTPException, UploadFile
from fastapi import FastAPI
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, Response, StreamingResponse
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel, Field

ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "data"
DB_PATH = DATA_DIR / "csd.db"
ORIGINALS_DIR = DATA_DIR / "originals"

app = FastAPI(title="CSD 備品管理系統", version="0.2.2")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://127.0.0.1:3000", "http://localhost:3000", "http://127.0.0.1:5173"],
    allow_methods=["*"], allow_headers=["*"],
)

TYPE_LABELS = {"rma": "RMA", "inventory": "庫存", "master": "料號主檔"}


def db() -> sqlite3.Connection:
    DATA_DIR.mkdir(exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def now_iso() -> str:
    return datetime.now().isoformat(timespec="seconds")


def version_code(created_at: str | None = None) -> str:
    value = datetime.fromisoformat(created_at) if created_at else datetime.now()
    return "V" + value.strftime("%y%m%d")


def clean_text(value: Any) -> str:
    return str(value or "").strip()


def as_num(value: Any) -> float:
    if value in (None, "", "-"):
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def strict_num(value: Any) -> float | None:
    if value in (None, "", "-"):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def as_date(value: Any) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, str):
        for fmt in (None, "%Y/%m/%d", "%Y-%m-%d"):
            try:
                if fmt:
                    return datetime.strptime(value.strip(), fmt).date().isoformat()
                return datetime.fromisoformat(value.strip().replace("/", "-")).date().isoformat()
            except ValueError:
                continue
    return None


def sheet_rows(ws: Any) -> list[tuple[Any, ...]]:
    return list(ws.iter_rows(values_only=True))


def json_rows(rows: list[sqlite3.Row]) -> list[dict[str, Any]]:
    return [dict(row) for row in rows]


def init_db() -> None:
    with db() as conn:
        conn.executescript("""
        CREATE TABLE IF NOT EXISTS import_batches (
          id INTEGER PRIMARY KEY, filename TEXT NOT NULL, created_at TEXT NOT NULL,
          summary_json TEXT NOT NULL, errors_json TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS parts (
          material_number TEXT PRIMARY KEY, site TEXT, model TEXT, pn_key TEXT, pn_key2 TEXT,
          target_months REAL NOT NULL DEFAULT 6, safety_months REAL NOT NULL DEFAULT 3,
          demand_override REAL, safety_override REAL, active INTEGER NOT NULL DEFAULT 1,
          inbound_qty REAL NOT NULL DEFAULT 0, imported_planned_qty REAL NOT NULL DEFAULT 0,
          notes TEXT, updated_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS stock_snapshots (
          id INTEGER PRIMARY KEY, batch_id INTEGER NOT NULL, material_number TEXT NOT NULL,
          warehouse TEXT NOT NULL, bin_location TEXT, quantity REAL NOT NULL,
          FOREIGN KEY(batch_id) REFERENCES import_batches(id)
        );
        CREATE INDEX IF NOT EXISTS stock_part_idx ON stock_snapshots(material_number);
        CREATE TABLE IF NOT EXISTS rma_records (
          id INTEGER PRIMARY KEY, batch_id INTEGER NOT NULL, service_date TEXT NOT NULL,
          customer TEXT, region TEXT, product_type TEXT, model_customer TEXT, model TEXT,
          serial_number TEXT, failure_classification TEXT, material_number TEXT,
          FOREIGN KEY(batch_id) REFERENCES import_batches(id)
        );
        CREATE INDEX IF NOT EXISTS rma_part_date_idx ON rma_records(material_number, service_date);
        CREATE TABLE IF NOT EXISTS shipments (
          id INTEGER PRIMARY KEY, batch_id INTEGER NOT NULL, material_number TEXT NOT NULL,
          shipment_year INTEGER, quantity REAL NOT NULL,
          FOREIGN KEY(batch_id) REFERENCES import_batches(id)
        );
        CREATE TABLE IF NOT EXISTS production_orders (
          id INTEGER PRIMARY KEY, material_number TEXT NOT NULL, suggested_qty REAL NOT NULL,
          confirmed_qty REAL, expected_date TEXT, status TEXT NOT NULL DEFAULT '草稿', notes TEXT,
          created_at TEXT NOT NULL, updated_at TEXT NOT NULL, deleted_at TEXT, deleted_by TEXT,
          FOREIGN KEY(material_number) REFERENCES parts(material_number)
        );
        CREATE TABLE IF NOT EXISTS import_versions (
          id INTEGER PRIMARY KEY, data_type TEXT NOT NULL, import_mode TEXT NOT NULL,
          filename TEXT NOT NULL, stored_filename TEXT, created_at TEXT NOT NULL, created_by TEXT NOT NULL,
          summary_json TEXT NOT NULL, warnings_json TEXT NOT NULL, errors_json TEXT NOT NULL,
          payload_json TEXT NOT NULL, scope_json TEXT NOT NULL DEFAULT '[]', restored_from_id INTEGER,
          batch_id INTEGER, version_code TEXT, FOREIGN KEY(restored_from_id) REFERENCES import_versions(id),
          FOREIGN KEY(batch_id) REFERENCES import_batches(id)
        );
        CREATE INDEX IF NOT EXISTS import_version_type_idx ON import_versions(data_type, id DESC);
        """)
        production_columns = {row[1] for row in conn.execute("PRAGMA table_info(production_orders)")}
        for name, definition in (("deleted_at", "TEXT"), ("deleted_by", "TEXT")):
            if name not in production_columns:
                conn.execute(f"ALTER TABLE production_orders ADD COLUMN {name} {definition}")
        version_columns = {row[1] for row in conn.execute("PRAGMA table_info(import_versions)")}
        if "version_code" not in version_columns:
            conn.execute("ALTER TABLE import_versions ADD COLUMN version_code TEXT")
        for row in conn.execute("SELECT id,created_at FROM import_versions WHERE version_code IS NULL OR version_code='' ").fetchall():
            conn.execute("UPDATE import_versions SET version_code=? WHERE id=?", (version_code(row["created_at"]), row["id"]))
        # Existing local trial data becomes a recoverable first version after this upgrade.
        has_version = conn.execute("SELECT 1 FROM import_versions LIMIT 1").fetchone()
        has_parts = conn.execute("SELECT 1 FROM parts LIMIT 1").fetchone()
        if not has_version and has_parts:
            created = now_iso()
            for data_type, payload in (("master", current_master(conn)), ("rma", current_rmas(conn)), ("inventory", current_stocks(conn))):
                if not payload:
                    continue
                batch = new_batch(conn, "既有本機資料（第一版）", created, {"records": len(payload)}, [])
                conn.execute("""INSERT INTO import_versions(data_type,import_mode,filename,stored_filename,created_at,created_by,summary_json,warnings_json,errors_json,payload_json,scope_json,batch_id,version_code)
                    VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)""", (
                    data_type, "baseline", "既有本機資料", None, created, "本機使用者",
                    json.dumps({"records": len(payload)}, ensure_ascii=False), "[]", "[]",
                    json.dumps(payload, ensure_ascii=False), json.dumps(sorted({x.get("warehouse", "") for x in payload}) if data_type == "inventory" else [], ensure_ascii=False), batch, version_code(created),
                ))


def new_batch(conn: sqlite3.Connection, filename: str, created: str, summary: dict[str, Any], messages: list[str]) -> int:
    cur = conn.execute(
        "INSERT INTO import_batches(filename,created_at,summary_json,errors_json) VALUES(?,?,?,?)",
        (filename, created, json.dumps(summary, ensure_ascii=False), json.dumps(messages, ensure_ascii=False)),
    )
    return int(cur.lastrowid)


def current_master(conn: sqlite3.Connection) -> list[dict[str, Any]]:
    rows = conn.execute("SELECT material_number,site,model,pn_key,pn_key2,inbound_qty,imported_planned_qty,notes,active FROM parts WHERE active=1 ORDER BY material_number").fetchall()
    return json_rows(rows)


def current_rmas(conn: sqlite3.Connection) -> list[dict[str, Any]]:
    return json_rows(conn.execute("""SELECT service_date,customer,region,product_type,model_customer,model,serial_number,failure_classification,material_number
        FROM rma_records ORDER BY id""").fetchall())


def current_stocks(conn: sqlite3.Connection) -> list[dict[str, Any]]:
    return json_rows(conn.execute("SELECT material_number,warehouse,bin_location,quantity FROM stock_snapshots ORDER BY id").fetchall())


def rma_key(item: dict[str, Any]) -> str:
    return "|".join(clean_text(item.get(key)).upper() for key in ("material_number", "service_date", "serial_number", "customer", "failure_classification"))


def headers_from_sheet(ws: Any, sheet_name: str) -> tuple[dict[str, int], list[str], list[str]]:
    rows = sheet_rows(ws)
    if not rows:
        return {}, [f"{sheet_name}：工作表沒有資料"], []
    headers = [clean_text(value) for value in rows[0]]
    present = [value for value in headers if value]
    duplicate = sorted(name for name, count in Counter(present).items() if count > 1)
    errors = [f"{sheet_name}：欄位「{name}」重複" for name in duplicate]
    return {name: index for index, name in enumerate(headers) if name}, errors, headers


def value_at(row: tuple[Any, ...], header: dict[str, int], name: str) -> Any:
    index = header.get(name)
    return row[index] if index is not None and index < len(row) else None


def parse_update_workbook(path: Path, data_type: str, mode: str) -> tuple[dict[str, Any], list[dict[str, Any]], list[str], list[str], list[str]]:
    expected_sheet = {"rma": "RMA", "inventory": "庫存快照", "master": "料號主檔"}.get(data_type)
    if not expected_sheet:
        raise ValueError("無效的資料類型")
    workbook = load_workbook(path, read_only=True, data_only=True)
    if expected_sheet not in workbook.sheetnames:
        raise ValueError(f"缺少必要工作表：{expected_sheet}")
    ws = workbook[expected_sheet]
    header, errors, _ = headers_from_sheet(ws, expected_sheet)
    required = {
        "rma": ["RMA日期", "料號"],
        "inventory": ["倉庫", "料號", "數量"],
        "master": ["料號"],
    }[data_type]
    for name in required:
        if name not in header:
            errors.append(f"{expected_sheet}：缺少必填欄位「{name}」")
    if errors:
        return {"records": 0, "errors": len(errors), "warnings": 0}, [], errors, [], []

    items: list[dict[str, Any]] = []
    warnings: list[str] = []
    row_errors: list[str] = []
    known_parts: set[str]
    with db() as conn:
        known_parts = {row[0] for row in conn.execute("SELECT material_number FROM parts")}
    seen: set[str] = set()
    rows = sheet_rows(ws)
    for number, row in enumerate(rows[1:], start=2):
        if not any(value not in (None, "") for value in row):
            continue
        material = clean_text(value_at(row, header, "料號"))
        if not material:
            row_errors.append(f"{expected_sheet} 第 {number} 列：料號不可空白")
            continue
        if data_type != "master" and material not in known_parts:
            row_errors.append(f"{expected_sheet} 第 {number} 列：未知料號「{material}」")
            continue
        if data_type == "rma":
            service_date = as_date(value_at(row, header, "RMA日期"))
            if not service_date:
                row_errors.append(f"RMA 第 {number} 列：RMA日期格式無效")
                continue
            item = {
                "service_date": service_date, "material_number": material,
                "customer": clean_text(value_at(row, header, "客戶")), "region": clean_text(value_at(row, header, "地區")),
                "product_type": clean_text(value_at(row, header, "產品類別")), "model_customer": clean_text(value_at(row, header, "客戶機種")),
                "model": clean_text(value_at(row, header, "機種")), "serial_number": clean_text(value_at(row, header, "序號")),
                "failure_classification": clean_text(value_at(row, header, "故障分類")),
            }
            key = rma_key(item)
            if key in seen:
                warnings.append(f"RMA 第 {number} 列：與本檔其他資料重複，已略過")
                continue
            seen.add(key)
            if not item["serial_number"]:
                warnings.append(f"RMA 第 {number} 列：序號空白，重複判定準確度較低")
            items.append(item)
        elif data_type == "inventory":
            warehouse = clean_text(value_at(row, header, "倉庫"))
            quantity = strict_num(value_at(row, header, "數量"))
            if not warehouse:
                row_errors.append(f"庫存快照 第 {number} 列：倉庫不可空白")
                continue
            if quantity is None:
                row_errors.append(f"庫存快照 第 {number} 列：數量必須是數字")
                continue
            if quantity < 0:
                row_errors.append(f"庫存快照 第 {number} 列：不可輸入負庫存")
                continue
            if quantity == 0:
                warnings.append(f"庫存快照 第 {number} 列：數量為 0")
            items.append({"material_number": material, "warehouse": warehouse, "bin_location": clean_text(value_at(row, header, "儲位")) or None, "quantity": quantity})
        else:
            inbound = strict_num(value_at(row, header, "在途量"))
            planned = strict_num(value_at(row, header, "既有待投產量"))
            if inbound is not None and inbound < 0:
                row_errors.append(f"料號主檔 第 {number} 列：在途量不可為負數")
                continue
            if planned is not None and planned < 0:
                row_errors.append(f"料號主檔 第 {number} 列：既有待投產量不可為負數")
                continue
            if material in seen:
                row_errors.append(f"料號主檔 第 {number} 列：料號「{material}」重複")
                continue
            seen.add(material)
            items.append({
                "material_number": material, "site": clean_text(value_at(row, header, "Site")), "model": clean_text(value_at(row, header, "機種")),
                "pn_key": clean_text(value_at(row, header, "PN Key")), "pn_key2": clean_text(value_at(row, header, "PN Key 2")),
                "inbound_qty": inbound or 0, "imported_planned_qty": planned or 0, "notes": clean_text(value_at(row, header, "備註")), "active": 1,
            })

    if data_type == "rma" and mode == "incremental" and not row_errors:
        with db() as conn:
            existing = {rma_key(x) for x in current_rmas(conn)}
        retained = []
        for item in items:
            if rma_key(item) in existing:
                warnings.append(f"RMA：{item['material_number']} / {item['service_date']} 已存在，增量匯入時略過")
            else:
                retained.append(item)
        items = retained
    if data_type == "inventory" and mode == "partial" and items:
        warnings.append("指定倉庫更新只會取代本檔案出現的倉庫，其餘倉庫資料維持不變")
    summary = {"records": len(items), "errors": len(errors) + len(row_errors), "warnings": len(warnings), "warehouses": len({x.get('warehouse') for x in items if x.get('warehouse')})}
    return summary, items, errors + row_errors, warnings, sorted({x.get("warehouse", "") for x in items if x.get("warehouse")})


def safe_name(filename: str) -> str:
    return re.sub(r"[^A-Za-z0-9._-]", "_", Path(filename).name) or "upload.xlsx"


def store_original(path: Path, version_id: int, filename: str) -> str:
    ORIGINALS_DIR.mkdir(parents=True, exist_ok=True)
    stored = f"{version_id:06d}-{safe_name(filename)}"
    (ORIGINALS_DIR / stored).write_bytes(path.read_bytes())
    return stored


def apply_payload(conn: sqlite3.Connection, data_type: str, payload: list[dict[str, Any]], batch_id: int, scope: list[str]) -> None:
    if data_type == "rma":
        conn.execute("DELETE FROM rma_records")
        conn.executemany("""INSERT INTO rma_records(batch_id,service_date,customer,region,product_type,model_customer,model,serial_number,failure_classification,material_number)
            VALUES(:batch_id,:service_date,:customer,:region,:product_type,:model_customer,:model,:serial_number,:failure_classification,:material_number)""", [{**item, "batch_id": batch_id} for item in payload])
    elif data_type == "inventory":
        warehouses = scope or sorted({item["warehouse"] for item in payload})
        if warehouses:
            placeholders = ",".join("?" for _ in warehouses)
            conn.execute(f"DELETE FROM stock_snapshots WHERE warehouse IN ({placeholders})", warehouses)
        conn.executemany("INSERT INTO stock_snapshots(batch_id,material_number,warehouse,bin_location,quantity) VALUES(:batch_id,:material_number,:warehouse,:bin_location,:quantity)", [{**item, "batch_id": batch_id} for item in payload])
    elif data_type == "master":
        conn.execute("UPDATE parts SET active=0")
        for item in payload:
            conn.execute("""INSERT INTO parts(material_number,site,model,pn_key,pn_key2,inbound_qty,imported_planned_qty,notes,active,updated_at)
                VALUES(:material_number,:site,:model,:pn_key,:pn_key2,:inbound_qty,:imported_planned_qty,:notes,1,:updated_at)
                ON CONFLICT(material_number) DO UPDATE SET site=excluded.site,model=excluded.model,pn_key=excluded.pn_key,pn_key2=excluded.pn_key2,
                inbound_qty=excluded.inbound_qty,imported_planned_qty=excluded.imported_planned_qty,notes=excluded.notes,active=1,updated_at=excluded.updated_at""", {**item, "updated_at": now_iso()})


def save_version(data_type: str, mode: str, filename: str, file_path: Path, payload: list[dict[str, Any]], summary: dict[str, Any], warnings: list[str], errors: list[str], scope: list[str], created_by: str, restored_from_id: int | None = None) -> int:
    with db() as conn:
        created = now_iso()
        batch_id = new_batch(conn, filename, created, summary, warnings + errors)
        cur = conn.execute("""INSERT INTO import_versions(data_type,import_mode,filename,stored_filename,created_at,created_by,summary_json,warnings_json,errors_json,payload_json,scope_json,restored_from_id,batch_id,version_code)
            VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)""", (
            data_type, mode, filename, None, created, clean_text(created_by) or "本機使用者",
            json.dumps(summary, ensure_ascii=False), json.dumps(warnings, ensure_ascii=False), json.dumps(errors, ensure_ascii=False),
            json.dumps(payload, ensure_ascii=False), json.dumps(scope, ensure_ascii=False), restored_from_id, batch_id, version_code(created),
        ))
        version_id = int(cur.lastrowid)
        stored = store_original(file_path, version_id, filename) if file_path.exists() else None
        conn.execute("UPDATE import_versions SET stored_filename=? WHERE id=?", (stored, version_id))
        apply_payload(conn, data_type, payload, batch_id, scope)
        return version_id


def parse_workbook(path: Path) -> tuple[dict[str, Any], dict[str, list[dict[str, Any]]], list[str]]:
    workbook = load_workbook(path, read_only=True, data_only=True, keep_vba=path.suffix.lower() == ".xlsm")
    required = {"缺料表", "Overall battery list 2026", "CSD&客戶庫存", "總出貨"}
    missing = sorted(required - set(workbook.sheetnames))
    if missing:
        raise ValueError("缺少必要工作表：" + "、".join(missing))
    errors: list[str] = []
    parts: list[dict[str, Any]] = []
    for i, row in enumerate(sheet_rows(workbook["缺料表"])[1:], start=2):
        pin = clean_text(row[4] if len(row) > 4 else None)
        if not pin:
            continue
        if any(item["material_number"] == pin for item in parts):
            errors.append(f"缺料表第 {i} 列：料號 {pin} 重複，已略過重複列")
            continue
        parts.append({"material_number": pin, "site": clean_text(row[0]), "model": clean_text(row[1]), "pn_key": clean_text(row[2]), "pn_key2": clean_text(row[3]), "inbound_qty": max(0, as_num(row[20] if len(row) > 20 else None)), "imported_planned_qty": max(0, as_num(row[19] if len(row) > 19 else None)), "notes": clean_text(row[27] if len(row) > 27 else None), "active": 1})
    stocks: list[dict[str, Any]] = []
    stock_rows = sheet_rows(workbook["CSD&客戶庫存"])
    header = stock_rows[0] if stock_rows else []
    warehouse_columns = [(col, clean_text(label).replace(" STOCK", "")) for col, label in enumerate(header) if label and "STOCK" in str(label).upper()]
    for row in stock_rows[1:]:
        for col, warehouse in warehouse_columns:
            material = clean_text(row[col] if col < len(row) else None)
            qty = as_num(row[col + 1] if col + 1 < len(row) else None)
            if material:
                stocks.append({"material_number": material, "warehouse": warehouse, "bin_location": None, "quantity": qty})
    part_matchers = []
    for part in parts:
        keys = [re.sub(r"[^A-Z0-9]", "", part[key].upper()) for key in ("pn_key", "pn_key2", "model") if part[key]]
        part_matchers.append((part["material_number"], [key for key in keys if len(key) >= 4]))
    rmas: list[dict[str, Any]] = []
    for row in sheet_rows(workbook["Overall battery list 2026"])[1:]:
        service_date = as_date(row[1] if len(row) > 1 else None)
        if not service_date:
            continue
        model_customer, serial_number = clean_text(row[9] if len(row) > 9 else None), clean_text(row[10] if len(row) > 10 else None)
        search_text = re.sub(r"[^A-Z0-9]", "", (model_customer + " " + serial_number).upper())
        matched = next((material for material, keys in part_matchers if any(key in search_text for key in keys)), None)
        rmas.append({"service_date": service_date, "customer": clean_text(row[5] if len(row) > 5 else None), "region": clean_text(row[6] if len(row) > 6 else None), "product_type": clean_text(row[4] if len(row) > 4 else None), "model_customer": model_customer, "model": clean_text(row[12] if len(row) > 12 else None), "serial_number": serial_number, "failure_classification": clean_text(row[13] if len(row) > 13 else None), "material_number": matched})
    shipments: list[dict[str, Any]] = []
    for row in sheet_rows(workbook["總出貨"])[1:]:
        material = clean_text(row[0] if len(row) else None)
        if material:
            for idx, value in enumerate(row[2:], start=2020):
                if idx > 2035:
                    break
                if as_num(value):
                    shipments.append({"material_number": material, "shipment_year": idx, "quantity": as_num(value)})
    unmatched = sum(1 for item in rmas if not item["material_number"])
    summary = {"parts": len(parts), "stock_rows": len(stocks), "rma_records": len(rmas), "matched_rma_records": len(rmas) - unmatched, "unmatched_rma_records": unmatched, "shipment_rows": len(shipments), "errors": len(errors), "latest_rma_date": max((x["service_date"] for x in rmas), default=None)}
    return summary, {"parts": parts, "stocks": stocks, "rmas": rmas, "shipments": shipments}, errors


def save_import(filename: str, payload: dict[str, list[dict[str, Any]]], summary: dict[str, Any], errors: list[str], file_path: Path | None = None) -> int:
    init_db()
    with db() as conn:
        created = now_iso()
        batch_id = new_batch(conn, filename, created, summary, errors)
        conn.execute("DELETE FROM stock_snapshots")
        conn.execute("DELETE FROM rma_records")
        conn.execute("DELETE FROM shipments")
        apply_payload(conn, "master", payload["parts"], batch_id, [])
        apply_payload(conn, "inventory", payload["stocks"], batch_id, sorted({x["warehouse"] for x in payload["stocks"]}))
        conn.executemany("""INSERT INTO rma_records(batch_id,service_date,customer,region,product_type,model_customer,model,serial_number,failure_classification,material_number)
            VALUES(:batch_id,:service_date,:customer,:region,:product_type,:model_customer,:model,:serial_number,:failure_classification,:material_number)""", [{**x, "batch_id": batch_id} for x in payload["rmas"]])
        conn.executemany("INSERT INTO shipments(batch_id,material_number,shipment_year,quantity) VALUES(:batch_id,:material_number,:shipment_year,:quantity)", [{**x, "batch_id": batch_id} for x in payload["shipments"]])
    # Store individual snapshots for later restore. This happens after the complete bundle is live.
    source = file_path or Path("/nonexistent")
    for data_type, records, scope in (("master", payload["parts"], []), ("rma", payload["rmas"], []), ("inventory", payload["stocks"], sorted({x["warehouse"] for x in payload["stocks"]}))):
        save_version(data_type, "initial", filename, source, records, {"records": len(records), "source": "完整初始活頁簿"}, errors, [], scope, "本機使用者")
    return batch_id


def month_floor(dt: date, months: int) -> date:
    year, month = dt.year, dt.month - months
    while month <= 0:
        year -= 1
        month += 12
    return date(year, month, 1)


def planning_rows(search: str = "", site: str = "", only_shortage: bool = False) -> list[dict[str, Any]]:
    init_db()
    with db() as conn:
        rows = conn.execute("SELECT * FROM parts WHERE active=1 ORDER BY site, material_number").fetchall()
        latest = conn.execute("SELECT MAX(service_date) AS max_date FROM rma_records").fetchone()["max_date"]
        anchor = date.fromisoformat(latest) if latest else date.today()
        stock_rows = conn.execute("SELECT material_number, warehouse, SUM(quantity) qty FROM stock_snapshots GROUP BY material_number, warehouse").fetchall()
        stocks: dict[str, dict[str, float]] = {}
        for row in stock_rows:
            stocks.setdefault(row["material_number"], {})[row["warehouse"]] = float(row["qty"])
        rma_dates: dict[str, list[date]] = {}
        for row in conn.execute("SELECT material_number, service_date FROM rma_records WHERE material_number IS NOT NULL").fetchall():
            rma_dates.setdefault(row["material_number"], []).append(date.fromisoformat(row["service_date"]))
        production = {row["material_number"]: float(row["qty"] or 0) for row in conn.execute("SELECT material_number, SUM(COALESCE(confirmed_qty,suggested_qty)) qty FROM production_orders WHERE deleted_at IS NULL AND status IN ('已提出','生產中') GROUP BY material_number").fetchall()}
        result = []
        for row in rows:
            part = dict(row)
            if search and search.lower() not in " ".join(str(part[key] or "") for key in ("material_number", "site", "model")).lower():
                continue
            if site and part["site"] != site:
                continue
            dates = rma_dates.get(part["material_number"], [])
            counts = {months: sum(1 for value in dates if value >= month_floor(anchor, months)) for months in (1, 3, 6, 12)}
            derived = counts[1] * .4 + counts[3] / 3 * .3 + counts[6] / 6 * .2 + counts[12] / 12 * .1
            demand = float(part["demand_override"]) if part["demand_override"] is not None else derived
            target = demand * float(part["target_months"])
            safety = float(part["safety_override"]) if part["safety_override"] is not None else demand * float(part["safety_months"])
            by_warehouse = stocks.get(part["material_number"], {})
            csd = sum(qty for name, qty in by_warehouse.items() if name.upper().startswith("CSD"))
            overseas = sum(qty for name, qty in by_warehouse.items() if not name.upper().startswith("CSD"))
            pending = float(part["imported_planned_qty"]) + production.get(part["material_number"], 0)
            available = csd + float(part["inbound_qty"]) + pending
            shortage, suggested = max(0, safety - available), max(0, target - available)
            result.append({**part, "rma_1m": counts[1], "rma_3m": counts[3], "rma_6m": counts[6], "rma_12m": counts[12], "as_of_date": anchor.isoformat(), "monthly_demand": round(demand, 2), "target_stock": round(target, 2), "safety_stock": round(safety, 2), "csd_stock": csd, "overseas_stock": overseas, "in_transit": float(part["inbound_qty"]), "pending_production": pending, "available_csd": available, "shortage_qty": round(shortage, 2), "suggested_production": round(suggested, 2), "priority_score": round(shortage / safety if safety else 0, 3)})
        return sorted([item for item in result if not only_shortage or item["shortage_qty"] > 0], key=lambda item: (item["shortage_qty"], item["priority_score"], item["monthly_demand"]), reverse=True)


class PartPatch(BaseModel):
    target_months: float | None = Field(default=None, ge=0)
    safety_months: float | None = Field(default=None, ge=0)
    demand_override: float | None = Field(default=None, ge=0)
    safety_override: float | None = Field(default=None, ge=0)
    notes: str | None = None


class ProductionInput(BaseModel):
    material_number: str
    suggested_qty: float = Field(ge=0)
    confirmed_qty: float | None = Field(default=None, ge=0)
    expected_date: str | None = None
    status: str = "草稿"
    notes: str | None = None


class ProductionPatch(BaseModel):
    confirmed_qty: float | None = Field(default=None, ge=0)
    expected_date: str | None = None
    status: str | None = None
    notes: str | None = None


def normalize_production_record(status: str, expected_date: str | None, notes: str | None) -> tuple[str | None, str | None]:
    if status == "草稿":
        return expected_date, notes
    if not expected_date or not as_date(expected_date):
        raise HTTPException(400, "草稿以外的狀態必須填寫日期")
    if not clean_text(notes):
        raise HTTPException(400, "草稿以外的狀態必須填寫備註")
    normalized_date = as_date(expected_date)
    body = re.sub(r"^\[\d{4}-\d{2}-\d{2}\]\s*", "", clean_text(notes))
    return normalized_date, f"[{normalized_date}] {body}"


@app.on_event("startup")
def startup() -> None:
    init_db()


@app.get("/api/health")
def health() -> dict[str, str]:
    return {"status": "ok"}


@app.post("/api/import/preview")
async def preview_import(file: UploadFile = File(...)) -> dict[str, Any]:
    if not file.filename or Path(file.filename).suffix.lower() not in (".xlsx", ".xlsm"):
        raise HTTPException(400, "請上傳 .xlsx 或 .xlsm 檔案")
    temp = DATA_DIR / f"preview-{datetime.now().timestamp()}-{safe_name(file.filename)}"
    temp.write_bytes(await file.read())
    try:
        summary, _, errors = parse_workbook(temp)
        return {"filename": file.filename, "summary": summary, "errors": errors[:50], "warnings": errors[:50]}
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc
    finally:
        temp.unlink(missing_ok=True)


@app.post("/api/import/commit")
async def commit_import(file: UploadFile = File(...)) -> dict[str, Any]:
    if not file.filename or Path(file.filename).suffix.lower() not in (".xlsx", ".xlsm"):
        raise HTTPException(400, "請上傳 .xlsx 或 .xlsm 檔案")
    temp = DATA_DIR / f"import-{datetime.now().timestamp()}-{safe_name(file.filename)}"
    temp.write_bytes(await file.read())
    try:
        summary, payload, errors = parse_workbook(temp)
        batch_id = save_import(file.filename, payload, summary, errors, temp)
        return {"batch_id": batch_id, "summary": summary, "errors": errors[:50]}
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc
    finally:
        temp.unlink(missing_ok=True)


@app.post("/api/data/preview")
async def preview_data(data_type: str = Form(...), mode: str = Form("full"), file: UploadFile = File(...)) -> dict[str, Any]:
    if data_type not in ("rma", "inventory", "master") or not file.filename or Path(file.filename).suffix.lower() != ".xlsx":
        raise HTTPException(400, "請使用系統範本上傳 .xlsx 檔案")
    if data_type == "rma" and mode not in ("full", "incremental"):
        raise HTTPException(400, "RMA 匯入方式無效")
    if data_type == "inventory" and mode not in ("full", "partial"):
        raise HTTPException(400, "庫存匯入方式無效")
    if data_type == "master" and mode != "full":
        raise HTTPException(400, "料號主檔僅支援完整更新")
    temp = DATA_DIR / f"preview-{datetime.now().timestamp()}-{safe_name(file.filename)}"
    temp.write_bytes(await file.read())
    try:
        summary, _, errors, warnings, scope = parse_update_workbook(temp, data_type, mode)
        return {"filename": file.filename, "data_type": data_type, "mode": mode, "summary": summary, "errors": errors[:100], "warnings": warnings[:100], "scope": scope, "can_commit": not errors}
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc
    finally:
        temp.unlink(missing_ok=True)


@app.post("/api/data/commit")
async def commit_data(data_type: str = Form(...), mode: str = Form("full"), operator: str = Form("本機使用者"), file: UploadFile = File(...)) -> dict[str, Any]:
    if data_type not in ("rma", "inventory", "master") or not file.filename or Path(file.filename).suffix.lower() != ".xlsx":
        raise HTTPException(400, "請使用系統範本上傳 .xlsx 檔案")
    if (data_type == "rma" and mode not in ("full", "incremental")) or (data_type == "inventory" and mode not in ("full", "partial")) or (data_type == "master" and mode != "full"):
        raise HTTPException(400, "匯入方式無效")
    temp = DATA_DIR / f"update-{datetime.now().timestamp()}-{safe_name(file.filename)}"
    temp.write_bytes(await file.read())
    try:
        summary, incoming, errors, warnings, scope = parse_update_workbook(temp, data_type, mode)
        if errors:
            raise HTTPException(400, {"message": "檔案檢查未通過", "errors": errors[:100]})
        with db() as conn:
            if data_type == "rma":
                payload = incoming if mode == "full" else current_rmas(conn) + incoming
                # A full source may contain accidental duplicates; keep the first record and preserve a warning.
                deduped: list[dict[str, Any]] = []
                seen: set[str] = set()
                for item in payload:
                    if rma_key(item) in seen:
                        warnings.append(f"RMA：重複資料已略過（{item['material_number']} / {item['service_date']}）")
                    else:
                        seen.add(rma_key(item)); deduped.append(item)
                payload = deduped
                actual_scope: list[str] = []
            elif data_type == "inventory":
                payload, actual_scope = incoming, scope
                if mode == "full":
                    actual_scope = sorted({item["warehouse"] for item in incoming})
            else:
                payload, actual_scope = incoming, []
        version_id = save_version(data_type, mode, file.filename, temp, payload, summary, warnings, [], actual_scope, operator)
        return {"version_id": version_id, "summary": {**summary, "stored_records": len(payload)}, "warnings": warnings[:100]}
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc
    finally:
        temp.unlink(missing_ok=True)


def template_bytes(data_type: str) -> bytes:
    specs = {
        "rma": ("RMA", ["RMA日期", "料號", "客戶", "地區", "產品類別", "客戶機種", "機種", "序號", "故障分類"]),
        "inventory": ("庫存快照", ["倉庫", "料號", "數量", "儲位"]),
        "master": ("料號主檔", ["料號", "Site", "機種", "PN Key", "PN Key 2", "在途量", "既有待投產量", "備註"]),
    }
    if data_type not in specs:
        raise HTTPException(404, "找不到範本")
    book = Workbook(); sheet = book.active; sheet.title, headers = specs[data_type]
    sheet.append(headers)
    sheet.freeze_panes = "A2"
    for cell in sheet[1]:
        cell.font = cell.font.copy(bold=True)
    for idx, name in enumerate(headers, start=1):
        sheet.column_dimensions[chr(64 + idx)].width = max(13, len(name) * 2 + 4)
    output = io.BytesIO(); book.save(output)
    return output.getvalue()


@app.get("/api/data/templates/{data_type}")
def download_template(data_type: str) -> Response:
    return Response(template_bytes(data_type), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename=csd-{data_type}-template.xlsx"})


@app.get("/api/data/versions")
def list_versions(search: str = "", data_type: str = "") -> dict[str, Any]:
    init_db()
    with db() as conn:
        clauses, values = [], []
        if data_type in ("rma", "inventory", "master"):
            clauses.append("data_type=?"); values.append(data_type)
        if search.strip():
            clauses.append("(version_code LIKE ? OR filename LIKE ? OR created_by LIKE ?)")
            values.extend([f"%{search.strip()}%"] * 3)
        where = f"WHERE {' AND '.join(clauses)}" if clauses else ""
        rows = conn.execute(f"SELECT id,data_type,import_mode,filename,stored_filename,created_at,created_by,summary_json,warnings_json,errors_json,scope_json,restored_from_id,version_code FROM import_versions {where} ORDER BY id DESC", values).fetchall()
    items = []
    for row in rows:
        item = dict(row)
        item["label"] = TYPE_LABELS.get(item["data_type"], item["data_type"])
        item["summary"] = json.loads(item.pop("summary_json"))
        item["warnings"] = json.loads(item.pop("warnings_json"))
        item["errors"] = json.loads(item.pop("errors_json"))
        item["scope"] = json.loads(item.pop("scope_json"))
        item["has_original"] = bool(item["stored_filename"] and (ORIGINALS_DIR / item["stored_filename"]).exists())
        item["has_data_download"] = True
        items.append(item)
    return {"items": items}


@app.get("/api/data/versions/{version_id}/file")
def download_original(version_id: int) -> FileResponse:
    with db() as conn:
        row = conn.execute("SELECT filename,stored_filename FROM import_versions WHERE id=?", (version_id,)).fetchone()
    if not row or not row["stored_filename"] or not (ORIGINALS_DIR / row["stored_filename"]).exists():
        raise HTTPException(404, "此版本沒有可下載的原始檔")
    return FileResponse(ORIGINALS_DIR / row["stored_filename"], filename=row["filename"], media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


def version_payload_workbook(data_type: str, payload: list[dict[str, Any]]) -> bytes:
    book = Workbook(); sheet = book.active
    specs = {
        "rma": ("RMA", ["RMA日期", "料號", "客戶", "地區", "產品類別", "客戶機種", "機種", "序號", "故障分類"]),
        "inventory": ("庫存快照", ["倉庫", "料號", "數量", "儲位"]),
        "master": ("料號主檔", ["料號", "Site", "機種", "PN Key", "PN Key 2", "在途量", "既有待投產量", "備註"]),
    }
    if data_type not in specs:
        raise HTTPException(404, "此版本沒有可下載的資料")
    sheet.title, headers = specs[data_type]; sheet.append(headers); sheet.freeze_panes = "A2"
    for item in payload:
        if data_type == "rma": sheet.append([item.get("service_date"), item.get("material_number"), item.get("customer"), item.get("region"), item.get("product_type"), item.get("model_customer"), item.get("model"), item.get("serial_number"), item.get("failure_classification")])
        elif data_type == "inventory": sheet.append([item.get("warehouse"), item.get("material_number"), item.get("quantity"), item.get("bin_location")])
        else: sheet.append([item.get("material_number"), item.get("site"), item.get("model"), item.get("pn_key"), item.get("pn_key2"), item.get("inbound_qty"), item.get("imported_planned_qty"), item.get("notes")])
    output = io.BytesIO(); book.save(output)
    return output.getvalue()


@app.get("/api/data/versions/{version_id}/data")
def download_version_data(version_id: int) -> Response:
    with db() as conn:
        row = conn.execute("SELECT data_type,payload_json,version_code FROM import_versions WHERE id=?", (version_id,)).fetchone()
    if not row:
        raise HTTPException(404, "找不到版本")
    content = version_payload_workbook(row["data_type"], json.loads(row["payload_json"]))
    name = f"{row['version_code']}-{row['data_type']}-data.xlsx"
    return Response(content, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename={name}"})


@app.post("/api/data/versions/{version_id}/restore")
def restore_version(version_id: int, operator: str = Form("本機使用者")) -> dict[str, Any]:
    with db() as conn:
        source = conn.execute("SELECT * FROM import_versions WHERE id=?", (version_id,)).fetchone()
    if not source:
        raise HTTPException(404, "找不到版本")
    source_dict = dict(source)
    data_type, payload = source_dict["data_type"], json.loads(source_dict["payload_json"])
    scope = json.loads(source_dict["scope_json"])
    if data_type not in ("rma", "inventory", "master"):
        raise HTTPException(400, "此版本不可還原")
    summary = {"records": len(payload), "restored_from": version_id, "source_filename": source_dict["filename"]}
    placeholder = Path("/nonexistent")
    restored = save_version(data_type, "restore", f"回復 {source_dict['version_code']}－{source_dict['filename']}", placeholder, payload, summary, [f"已回復自版本 {source_dict['version_code']}"], [], scope, operator, version_id)
    return {"version_id": restored, "message": f"已回復 {TYPE_LABELS[data_type]} 版本 {source_dict['version_code']}"}


@app.get("/api/data/export/current.zip")
def export_current_data() -> Response:
    with db() as conn:
        data = {"rma": current_rmas(conn), "inventory": current_stocks(conn), "master": current_master(conn)}
    output = io.BytesIO()
    with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as archive:
        for kind, rows in data.items():
            book = Workbook(); sheet = book.active
            spec_headers = {"rma": ["RMA日期", "料號", "客戶", "地區", "產品類別", "客戶機種", "機種", "序號", "故障分類"], "inventory": ["倉庫", "料號", "數量", "儲位"], "master": ["料號", "Site", "機種", "PN Key", "PN Key 2", "在途量", "既有待投產量", "備註"]}[kind]
            sheet.title = {"rma": "RMA", "inventory": "庫存快照", "master": "料號主檔"}[kind]; sheet.append(spec_headers)
            for row in rows:
                if kind == "rma": sheet.append([row.get("service_date"), row.get("material_number"), row.get("customer"), row.get("region"), row.get("product_type"), row.get("model_customer"), row.get("model"), row.get("serial_number"), row.get("failure_classification")])
                elif kind == "inventory": sheet.append([row.get("warehouse"), row.get("material_number"), row.get("quantity"), row.get("bin_location")])
                else: sheet.append([row.get("material_number"), row.get("site"), row.get("model"), row.get("pn_key"), row.get("pn_key2"), row.get("inbound_qty"), row.get("imported_planned_qty"), row.get("notes")])
            book_file = io.BytesIO(); book.save(book_file); archive.writestr(f"csd-current-{kind}.xlsx", book_file.getvalue())
        archive.writestr("README.txt", f"CSD 目前有效資料\n匯出時間：{now_iso()}\n")
    return Response(output.getvalue(), media_type="application/zip", headers={"Content-Disposition": "attachment; filename=csd-current-data.zip"})


@app.get("/api/dashboard")
def dashboard() -> dict[str, Any]:
    rows = planning_rows()
    with db() as conn:
        recent = conn.execute("SELECT filename,created_at,data_type,version_code FROM import_versions ORDER BY id DESC LIMIT 1").fetchone()
        production = conn.execute("SELECT * FROM production_orders WHERE deleted_at IS NULL ORDER BY updated_at DESC LIMIT 8").fetchall()
    return {"metrics": {"part_count": len(rows), "shortage_part_count": sum(1 for item in rows if item["shortage_qty"] > 0), "shortage_qty": round(sum(item["shortage_qty"] for item in rows), 1), "in_transit_qty": round(sum(item["in_transit"] for item in rows), 1), "pending_production_qty": round(sum(item["pending_production"] for item in rows), 1)}, "priority_parts": rows[:8], "recent_import": dict(recent) if recent else None, "production_orders": json_rows(production)}


@app.get("/api/parts")
def list_parts(search: str = "", site: str = "", only_shortage: bool = False) -> dict[str, Any]:
    rows = planning_rows(search, site, only_shortage)
    return {"items": rows, "sites": sorted({item["site"] for item in rows if item["site"]}), "as_of_date": rows[0]["as_of_date"] if rows else None}


@app.patch("/api/parts/{material_number}")
def patch_part(material_number: str, changes: PartPatch) -> dict[str, Any]:
    values = changes.model_dump(exclude_unset=True)
    if not values:
        raise HTTPException(400, "沒有可更新的欄位")
    values["updated_at"] = now_iso()
    with db() as conn:
        cur = conn.execute(f"UPDATE parts SET {', '.join(f'{key}=:{key}' for key in values)} WHERE material_number=:material_number", {**values, "material_number": material_number})
        if not cur.rowcount:
            raise HTTPException(404, "找不到料號")
    return {"ok": True}


@app.get("/api/production-orders")
def production_orders() -> list[dict[str, Any]]:
    with db() as conn:
        return json_rows(conn.execute("SELECT * FROM production_orders WHERE deleted_at IS NULL ORDER BY updated_at DESC").fetchall())


@app.get("/api/production-orders/deleted")
def deleted_production_orders() -> list[dict[str, Any]]:
    with db() as conn:
        return json_rows(conn.execute("SELECT * FROM production_orders WHERE deleted_at IS NOT NULL ORDER BY deleted_at DESC").fetchall())


@app.post("/api/production-orders")
def create_production_order(item: ProductionInput) -> dict[str, Any]:
    if item.status not in ("草稿", "已提出", "生產中", "已完成", "取消"):
        raise HTTPException(400, "無效的狀態")
    expected_date, notes = normalize_production_record(item.status, item.expected_date, item.notes)
    with db() as conn:
        if not conn.execute("SELECT 1 FROM parts WHERE material_number=?", (item.material_number,)).fetchone():
            raise HTTPException(404, "找不到料號")
        ts = now_iso(); cur = conn.execute("INSERT INTO production_orders(material_number,suggested_qty,confirmed_qty,expected_date,status,notes,created_at,updated_at) VALUES(?,?,?,?,?,?,?,?)", (item.material_number, item.suggested_qty, item.confirmed_qty, expected_date, item.status, notes, ts, ts))
    return {"id": cur.lastrowid}


@app.patch("/api/production-orders/{order_id}")
def patch_production_order(order_id: int, changes: ProductionPatch) -> dict[str, Any]:
    values = changes.model_dump(exclude_unset=True)
    if values.get("status") and values["status"] not in ("草稿", "已提出", "生產中", "已完成", "取消"):
        raise HTTPException(400, "無效的狀態")
    with db() as conn:
        current = conn.execute("SELECT * FROM production_orders WHERE id=? AND deleted_at IS NULL", (order_id,)).fetchone()
        if not current:
            raise HTTPException(404, "找不到生產需求")
        status = values.get("status", current["status"])
        expected_date = values.get("expected_date", current["expected_date"])
        notes = values.get("notes", current["notes"])
        expected_date, notes = normalize_production_record(status, expected_date, notes)
        values.update({"status": status, "expected_date": expected_date, "notes": notes, "updated_at": now_iso()})
        conn.execute("UPDATE production_orders SET " + ", ".join(f"{key}=:{key}" for key in values) + " WHERE id=:id", {**values, "id": order_id})
    return {"ok": True}


@app.delete("/api/production-orders/{order_id}")
def delete_production_order(order_id: int) -> dict[str, Any]:
    with db() as conn:
        cur = conn.execute("UPDATE production_orders SET deleted_at=?, deleted_by=?, updated_at=? WHERE id=? AND deleted_at IS NULL", (now_iso(), "本機使用者", now_iso(), order_id))
        if not cur.rowcount:
            raise HTTPException(404, "找不到生產需求")
    return {"ok": True}


@app.post("/api/production-orders/{order_id}/restore")
def restore_production_order(order_id: int) -> dict[str, Any]:
    with db() as conn:
        cur = conn.execute("UPDATE production_orders SET deleted_at=NULL, deleted_by=NULL, updated_at=? WHERE id=? AND deleted_at IS NOT NULL", (now_iso(), order_id))
        if not cur.rowcount:
            raise HTTPException(404, "找不到已刪除的生產需求")
    return {"ok": True}


@app.get("/api/export/shortages.csv")
def export_shortages() -> StreamingResponse:
    rows = planning_rows(only_shortage=True); buffer = io.StringIO(); writer = csv.writer(buffer)
    writer.writerow(["料號", "Site", "機種", "1M RMA", "3M RMA", "6M RMA", "12M RMA", "每月需求", "CSD庫存", "海外庫存", "在途", "待投產", "安全庫存", "缺料量", "建議生產量"])
    for item in rows:
        writer.writerow([item["material_number"], item["site"], item["model"], item["rma_1m"], item["rma_3m"], item["rma_6m"], item["rma_12m"], item["monthly_demand"], item["csd_stock"], item["overseas_stock"], item["in_transit"], item["pending_production"], item["safety_stock"], item["shortage_qty"], item["suggested_production"]])
    return StreamingResponse(iter(["\ufeff" + buffer.getvalue()]), media_type="text/csv; charset=utf-8", headers={"Content-Disposition": "attachment; filename=csd-shortages.csv"})


if __name__ == "__main__":
    init_db()
    if len(sys.argv) == 3 and sys.argv[1] == "--import":
        source = Path(sys.argv[2]); summary, payload, errors = parse_workbook(source); batch = save_import(source.name, payload, summary, errors, source)
        print(json.dumps({"batch_id": batch, **summary}, ensure_ascii=False))
    else:
        import uvicorn
        uvicorn.run(app, host="127.0.0.1", port=8000)
